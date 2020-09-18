# -*- coding: utf-8 -*-
"""
# 安装依赖

pip install pandas openpyxl xlrd
"""
import os
import re
import glob
import random
import string

from datetime import datetime

import pandas as pd
import openpyxl

from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border


def random_string(length=5):
    """生成随机指定长度的字符串

    :return 随机字符串
    :rtype str
    """
    strings = string.ascii_lowercase + string.ascii_uppercase + string.digits
    return "".join([strings[random.randint(0, len(strings) - 1)] for _ in
                    range(length)])


def random_number(length=5):
    """随机生成指定长度的数字

    :return 随机数字
    :rtype int
    """
    strings = string.digits
    s = "".join([str(random.randint(1, 9))] +
                [strings[random.randint(0,
                                        len(strings) - 1)] for _ in range(length - 1)])
    return int(s)


def read_csv():
    os.remove("/tmp/coupon-logs.csv")
    csv_list = glob.glob('/tmp/coupon-log/*/*.csv')
    print(csv_list[0])

    df = None
    for index, csv_path in enumerate(csv_list):
        print(index + 1, csv_path)
        temp_df = pd.read_csv(csv_path)
        df = pd.concat([df, temp_df], axis=0, ignore_index=True)

    # 替换敏感信息
    for content in df["content"]:
        df["content"] = re.sub(r'"session":\s*"\w+"', '"session": "xxx"', content)

    # 删除重复行
    df = df.drop_duplicates()

    # 删除不需要的列
    df = df.drop(['__tag__:__client_ip__', '__tag__:__hostname__',
                  '__tag__:__receive_time__', '__tag__:_node_ip_', '__tag__:_node_name_',
                  '__topic__', '_container_ip_', '_container_name_', '_image_name_',
                  '_namespace_', '_pod_name_', '_pod_uid_', '_source_'], axis=1)

    # 按列排序
    df = df.sort_values(by=["_time_"])
    df = df.reset_index(drop=True)
    df.to_csv("/tmp/coupon-logs.csv")


def write_excel():
    file_path = "/tmp/test.xlsx"

    # 删除文件
    if os.path.exists(file_path):
        os.remove(file_path)

    # 设置列名
    columns = ["clientId", "map1", "map2", "map3", "map4"]

    # 创建数据结构
    df = pd.DataFrame(data=[], columns=columns)

    # 循环生成数据
    for i in range(10):
        # 生成每列数据
        client_id = random_string()
        map1 = random.choices(["vip1", "vip2"])
        map2 = random.choices(["家装", "车贷", "ETC"])
        map3 = random.choices(["渠道1", "渠道2", "渠道3"])
        map4 = random.choices(["支行1", "支行2", "支行3"])
        tmp_data = [client_id, map1, map2, map3, map4]

        # 生成新的数据结构
        tmp_df = pd.DataFrame(data=[tmp_data], columns=columns)

        # 追加
        df = df.append(tmp_df, ignore_index=True)

    # 保存到文件中
    df.to_excel(file_path, sheet_name="表格名字", index=False)


def get_denominations(value):
    if value is not None and str(value) != "nan" and str(value).startswith("减"):
        return int(value[1:]) * 100
    return None


def check_create_time(creat_time):
    return datetime(2020, 9, 16) < datetime.strptime(creat_time, "%Y-%m-%d %H:%M:%S") < datetime(2020, 9, 17, 23, 59,
                                                                                                 59, 999)


def pandas_update_excel(coupons=None):
    coupons = coupons or []
    excel_path = "/Users/seekplum/Downloads/优惠券列表.xlsx"
    file_path = "/tmp/pandas-test.xlsx"
    column_list = [3, 12, 19]

    df = pd.read_excel(excel_path)
    df_list = df.values.tolist()
    for index, item in enumerate(df_list):
        for col in column_list:
            denominations = get_denominations(item[col])
            if denominations is None:
                continue
            for coupon in coupons:
                if not (coupon["denominations"] == denominations and check_create_time(coupon["creat_time"])):
                    continue
                old = df.iloc[index, col + 1]
                old = f"{old}\n" if old and str(old) != "nan" else ""
                df.iloc[index, col + 1] = f'{old}{coupon["activity_url"]}'
    df.to_excel(file_path, sheet_name="表格名字", index=False)


def write_cell(sheet, column, new_column, coupons):
    col_values = list(sheet.columns)[column]
    for row, cell in enumerate(col_values):
        denominations = get_denominations(cell.value)
        if denominations is None:
            continue
        new_row = row + 1
        for coupon in coupons:
            if not (coupon["denominations"] == denominations and check_create_time(coupon["creat_time"])):
                continue

            old = sheet.cell(new_row, new_column).value
            old = f"{old}\n" if old and str(old) != "nan" else ""
            new_value = f'{old}{coupon["activity_url"]}'

            sheet.cell(new_row, new_column, new_value)


def openpyxl_update_excel(coupons=None):
    coupons = coupons or []
    excel_path = "/Users/seekplum/Downloads/优惠券列表.xlsx"
    file_path = "/tmp/openpyxl-test.xlsx"
    column_list = [3, 12, 19]

    book = openpyxl.load_workbook(excel_path)
    sheet = book.active
    title_row = 3
    for index, column in enumerate(column_list):
        new_column = column + 2
        sheet.insert_cols(new_column)
        # 设置列宽
        sheet.column_dimensions[get_column_letter(new_column)].width = 120

        # 设置标题
        sheet.cell(title_row, new_column, "领取链接")

        # 设置背景颜色
        title_cell = sheet.cell(title_row, column)
        title_fill = title_cell.fill
        title_border = title_cell.border
        cell = sheet.cell(title_row, new_column)
        cell.fill = PatternFill("solid", fgColor=title_fill.fgColor, bgColor=title_fill.bgColor)
        cell.border = Border(left=title_border.left, right=title_border.right, top=title_border.top,
                             bottom=title_border.bottom)

        write_cell(sheet, column, new_column, coupons)

    write_cell(sheet, 2, 5, coupons)

    book.save(file_path)


if __name__ == '__main__':
    read_csv()
    write_excel()
    coupons = []
    pandas_update_excel(coupons)
    openpyxl_update_excel(coupons)
